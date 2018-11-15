import os
from functools import reduce

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook
from selenium import webdriver

file_name = 'tmp.xlsx'
title = 'test0001'

limit = 20
page_start = 0
_type = 'movie'
url = 'https://movie.douban.com/explore#!type=' + _type \
      + '&tag=%E7%83%AD%E9%97%A8&sort=recommend&page_limit=' + str(limit) \
      + '&page_start=' + str(page_start)

browser = webdriver.Firefox()
webdriver.Chrome()

def spider():
    r = requests.get(url)
    content = BeautifulSoup(r.content, 'html.parser')
    print(content)


def spider2():
    browser.get(url)
    return BeautifulSoup(browser.page_source, 'html.parser')


def work():
    data_list = list()
    soup = spider2()
    # div class="list-wp"
    base_div = soup.find('div', attrs={'class': 'list-wp'})
    # a class="item"
    a_s = base_div.find_all('a', attrs={'class': 'item'})
    for item in a_s:
        title_score = item.find('p').get_text().strip()
        result_map = dict()
        result_map['score'] = title_score
        data_list.append(result_map)
    output_xls(data_list)


def output_xls(data_list):
    print(data_list)
    if not os.path.exists(file_name):
        wb = Workbook()
    else:
        wb = load_workbook(file_name)

    if title not in wb.sheetnames:
        work_sheet = wb.create_sheet(title=title)
    else:
        work_sheet = wb[title]
    # printf
    for row in range(0, len(data_list)):
        for i in range(0, len(data_list[row])):
            _ = work_sheet.cell(column=i + 1, row=row + 1, value="%s" % data_list[row]['score'])
    wb.save(filename=file_name)

if __name__ == '__main__':
    work()
    browser.close()
